import openpyxl, sys

# создаем новую рабочую книгу Excel и выбираем активный лист
wb = openpyxl.Workbook()
sheet =	wb.active

dimension = 6   # размерность таблицы умножения по умолчанию

param = None

if len(sys.argv) > 1:
   param = sys.argv[1]

if param:
   try:
       dimension = int(param)
   except ValueError as error:
       print("Ошибка параметра командной строки. Необходимо указать целое число")
       print("Размерность таблицы умножения принята равной", dimension)
   
# прописать 1-ую строку
for col in range(2, dimension+2):
    sheet.cell(row=1, column=col).value = col - 1

# прописать остальные dimension строк, начиная со 2-ой колонки - с формулами
for cur_row in range(2, dimension+2):
    for col in range(1, dimension+2):
        if col == 1: # прописать константу в 1-ую колонку
            sheet.cell(row=cur_row, column=1).value = cur_row - 1
        else:        # прописать вычисленное значение, начиная со 2-ой колонки
            sheet.cell(row=cur_row, column=col).value = (cur_row - 1) * (col - 1)

# сохраняем рабочую книгу Excel
wb.save('multiplication.xlsx')

print("Записан файл:", 'multiplication.xlsx')
input("\nДля выхода нажмите Enter.")
