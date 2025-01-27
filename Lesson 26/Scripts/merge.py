import openpyxl

# создаем новую рабочую книгу Excel и выбираем активный лист
wb = openpyxl.Workbook()
sheet = wb.active

# сливаем ячейки в заданных прямоугольных областях
sheet.merge_cells('A1:D3')
sheet['A1'] = '12 слитых вместе ячеек.'

sheet.merge_cells('C5:D5')
sheet['C5'] = '2 слитых ячейки.'

# сохраняем рабочую книгу Excel
wb.save('merged.xlsx')

print("Записан файл:", 'merged.xlsx')
input("\nДля выхода нажмите Enter.")
