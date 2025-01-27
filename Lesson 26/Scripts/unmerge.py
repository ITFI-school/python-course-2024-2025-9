import openpyxl

# открываем рабочую книгу Excel из файла и выбираем активный лист
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active

# разъединяем ячейки в заданных прямоугольных областях
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')

# сохраняем рабочую книгу Excel
wb.save('unmerged.xlsx')

print("Записан файл:", 'unmerged.xlsx')
input("\nДля выхода нажмите Enter.")
