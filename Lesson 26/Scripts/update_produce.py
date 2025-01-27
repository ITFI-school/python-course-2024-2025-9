# updateProduce.py - корректирует цены товаров в листе с продажами

import openpyxl

print("Загружаем рабочую книгу Excel...")

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Название товаров с их новыми ценами
PRICE_UPDATES = {'Чеснок': 3.07,
                 'Сельдерей': 1.19,
                 'Лимоны': 1.27}

print("Проход по списку товаров на листе...")

# Цикл через строки данных на листе и обновление цен заданных товаров
for row_num in range(2, sheet.max_row): # пропустить 1-ую строку с заголовком
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')

print("Записан файл:", 'updatedProduceSales.xlsx')
input("\nДля выхода нажмите Enter.")
